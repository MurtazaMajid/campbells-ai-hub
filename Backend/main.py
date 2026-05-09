"""
Campbell's Restaurant — AI Marketing System
FastAPI Backend v4.0 — Full with Analysis + Customer Profiling (Section 5)

Endpoints:
  GET  /                                  → health check
  POST /api/segment                       → customer segment
  POST /api/churn                         → churn probability + risk level
  POST /api/sentiment                     → ABSA triplets from review
  POST /api/generate-message              → enriched personalized SMS + email + app notification
  POST /api/full-pipeline                 → all of the above in one call (auto-fetches profile from DB)
  GET  /api/dashboard                     → full stats (live from Supabase)
  GET  /api/customers                     → paginated customer list
  GET  /api/messages-log                  → recent generated messages
  GET  /api/customer-profile/{id}         → full enriched profile for one customer
  GET  /api/analysis/kpis                 → all KPI numbers
  GET  /api/analysis/rfm                  → RFM scatter data
  GET  /api/analysis/churn-distribution   → churn histogram + by segment
  GET  /api/analysis/monthly-visits       → visits by month
  GET  /api/analysis/revenue              → revenue breakdown
  GET  /api/analysis/sentiment-breakdown  → ABSA charts data
  GET  /api/analysis/profiles             → Section 5 spending tier / time / food breakdowns
"""

from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional, List
import pandas as pd
import numpy as np
import pickle
import json
import re
import os
import ast
import time
import requests
import warnings
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from supabase import create_client, Client
from sklearn.preprocessing import StandardScaler
from sklearn.cluster import KMeans
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.multiclass import OneVsRestClassifier
from sklearn.preprocessing import MultiLabelBinarizer
from xgboost import XGBClassifier
from openpyxl import load_workbook
warnings.filterwarnings('ignore')

# Guard resend import — app must not crash if package is absent or key is missing
try:
    import resend as _resend_module
    _resend_available = True
except ImportError:
    _resend_module   = None
    _resend_available = False

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
GROQ_API_KEY    = os.getenv("GROQ_API_KEY",    "YOUR_GROQ_API_KEY_HERE")
SUPABASE_URL    = os.getenv("SUPABASE_URL",    "YOUR_SUPABASE_URL_HERE")
SUPABASE_KEY    = os.getenv("SUPABASE_KEY",    "YOUR_SUPABASE_ANON_KEY_HERE")
RESEND_API_KEY  = os.getenv("RESEND_API_KEY",  "YOUR_RESEND_API_KEY_HERE")
GROQ_MODEL      = "llama-3.3-70b-versatile"
DATA_DIR        = os.getenv("DATA_DIR", "/app")   # PATCHED: was "." — set DATA_DIR=/app in Render env
DISCOUNT_MAP    = {'High': 20, 'Medium': 15, 'Low': 10}

# Only configure resend if available AND key is real
if _resend_available and RESEND_API_KEY != "YOUR_RESEND_API_KEY_HERE":
    _resend_module.api_key = RESEND_API_KEY
    resend = _resend_module
else:
    resend = None

# ─────────────────────────────────────────────
# SUPABASE CLIENT  (initialised once at module level — thread-safe)
# ─────────────────────────────────────────────
_supabase_client: Client = None

def get_supabase() -> Client:
    global _supabase_client
    if _supabase_client is None:
        _supabase_client = create_client(SUPABASE_URL, SUPABASE_KEY)
    return _supabase_client

# ─────────────────────────────────────────────
# PYDANTIC SCHEMAS
# ─────────────────────────────────────────────
class CustomerRequest(BaseModel):
    recency         : float
    frequency       : int
    monetary        : float
    unique_items    : Optional[int]       = 1
    avg_order_val   : Optional[float]     = None
    avg_tip         : Optional[float]     = 0.0
    discount_used   : Optional[int]       = 0
    visits_nov      : Optional[int]       = 0
    visits_dec      : Optional[int]       = 0
    visits_jan      : Optional[int]       = 0
    days_since_first: Optional[int]       = None
    favorite_items  : Optional[List[str]] = []

class SentimentRequest(BaseModel):
    review: str

class MessageRequest(BaseModel):
    customer_id      : Optional[str]       = None
    segment          : str
    recency          : float
    frequency        : int
    monetary         : float
    risk_level       : str
    churn_probability: float
    favorite_items   : Optional[List[str]] = []
    aspects          : Optional[List[str]] = []
    sentiments       : Optional[List[str]] = []
    # Section 5 enriched profile fields
    spending_tier    : Optional[str]  = "Standard"
    time_preference  : Optional[str]  = "Evening"
    food_preference  : Optional[str]  = "varied"
    drink_vs_food    : Optional[str]  = "Mixed"
    favorite_modifier: Optional[str]  = None
    is_flight_lover  : Optional[bool] = False

class FullPipelineRequest(BaseModel):
    recency         : float
    frequency       : int
    monetary        : float
    unique_items    : Optional[int]       = 1
    avg_order_val   : Optional[float]     = None
    avg_tip         : Optional[float]     = 0.0
    discount_used   : Optional[int]       = 0
    visits_nov      : Optional[int]       = 0
    visits_dec      : Optional[int]       = 0
    visits_jan      : Optional[int]       = 0
    days_since_first: Optional[int]       = None
    favorite_items  : Optional[List[str]] = []
    review          : Optional[str]       = None
    customer_id     : Optional[str]       = None
    # Section 5 profile fields (auto-fetched from DB if customer_id given and these are None)
    spending_tier    : Optional[str]  = None
    time_preference  : Optional[str]  = None
    food_preference  : Optional[str]  = None
    drink_vs_food    : Optional[str]  = None
    favorite_modifier: Optional[str]  = None
    is_flight_lover  : Optional[bool] = None

# ─────────────────────────────────────────────
# MODEL STORE
# ─────────────────────────────────────────────
models = {}

def load_models():
    global models
    try:
        with open(f"{DATA_DIR}/kmeans_model.pkl",      'rb') as f: models['kmeans']      = pickle.load(f)
        with open(f"{DATA_DIR}/scaler.pkl",            'rb') as f: models['scaler']      = pickle.load(f)
        with open(f"{DATA_DIR}/cluster_map.pkl",       'rb') as f: models['cluster_map'] = pickle.load(f)
        with open(f"{DATA_DIR}/churn_model_tier2.pkl", 'rb') as f: models['churn']       = pickle.load(f)
        with open(f"{DATA_DIR}/churn_features.pkl",    'rb') as f: models['churn_feats'] = pickle.load(f)
        with open(f"{DATA_DIR}/aspect_model.pkl",      'rb') as f: models['aspect']      = pickle.load(f)
        with open(f"{DATA_DIR}/tfidf_aspect.pkl",      'rb') as f: models['tfidf_asp']   = pickle.load(f)
        with open(f"{DATA_DIR}/mlb.pkl",               'rb') as f: models['mlb']         = pickle.load(f)
        with open(f"{DATA_DIR}/sentiment_model.pkl",   'rb') as f: models['sentiment']   = pickle.load(f)
        with open(f"{DATA_DIR}/tfidf_sent.pkl",        'rb') as f: models['tfidf_sent']  = pickle.load(f)
        print("All models loaded successfully")
    except FileNotFoundError as e:
        print(f"Model file not found: {e}")

def load_menu():
    try:
        wb      = load_workbook(f"{DATA_DIR}/Campbell_Menu_Data_-_2.xlsx", read_only=True)
        ws      = wb.active
        rows    = list(ws.iter_rows(values_only=True))
        menu_df = pd.DataFrame(rows[1:], columns=rows[0])
        food_cats = ['Signature Flights','Brunch Food','Entrees','Desserts',
                     'Salads','Burgers & Sandwiches','Kids Menu',
                     'Sides Dinner','Sides Brunch','Weekly Specials']
        return menu_df[menu_df['Category'].isin(food_cats)].dropna(subset=['itemName','itemPrice'])
    except:
        return pd.DataFrame(columns=['itemName','itemPrice','Category'])

menu_df = pd.DataFrame()

# ─────────────────────────────────────────────
# SUPABASE TABLE CREATION SQL
# ─────────────────────────────────────────────
CREATE_CUSTOMERS_SQL = """
CREATE TABLE IF NOT EXISTS customers (
    id                TEXT PRIMARY KEY,
    segment           TEXT,
    recency           FLOAT,
    frequency         INTEGER,
    monetary          FLOAT,
    unique_items      INTEGER,
    avg_order_val     FLOAT,
    avg_tip           FLOAT,
    discount_used     INTEGER,
    visits_nov        INTEGER,
    visits_dec        INTEGER,
    visits_jan        INTEGER,
    days_since_first  INTEGER,
    churn_probability FLOAT,
    risk_level        TEXT,
    tier              TEXT,
    discount_offered  TEXT,
    spending_tier     TEXT DEFAULT 'Standard',
    time_preference   TEXT DEFAULT 'Evening',
    food_preference   TEXT DEFAULT 'varied',
    drink_vs_food     TEXT DEFAULT 'Mixed',
    favorite_modifier TEXT,
    is_flight_lover   BOOLEAN DEFAULT FALSE,
    favorite_items    TEXT,
    email             TEXT,
    email_sent        BOOLEAN DEFAULT FALSE,
    email_sent_at     TIMESTAMPTZ,
    created_at        TIMESTAMPTZ DEFAULT NOW(),
    updated_at        TIMESTAMPTZ DEFAULT NOW()
);"""

CREATE_MESSAGES_LOG_SQL = """
CREATE TABLE IF NOT EXISTS messages_log (
    id               BIGSERIAL PRIMARY KEY,
    customer_id      TEXT,
    segment          TEXT,
    risk_level       TEXT,
    discount_offered TEXT,
    spending_tier    TEXT,
    time_preference  TEXT,
    sms              TEXT,
    email_subject    TEXT,
    email_body       TEXT,
    app_notification TEXT,
    aspects          JSONB,
    sentiments       JSONB,
    created_at       TIMESTAMPTZ DEFAULT NOW()
);"""

CREATE_ABSA_SQL = """
CREATE TABLE IF NOT EXISTS absa_predictions (
    id               BIGSERIAL PRIMARY KEY,
    customer_id      TEXT,
    review           TEXT,
    aspects          TEXT,
    sentiments       TEXT,
    feature_opinion  TEXT,
    opinions         TEXT,
    triplets         JSONB,
    created_at       TIMESTAMPTZ DEFAULT NOW()
);"""

def ensure_tables_via_api():
    service_key = os.getenv("SUPABASE_SERVICE_KEY", "")
    project_ref = SUPABASE_URL.replace("https://", "").replace(".supabase.co", "")
    if not service_key:
        print("SUPABASE_SERVICE_KEY not set - skipping auto table creation.")
        return False
    all_sql = "\n".join([CREATE_CUSTOMERS_SQL, CREATE_MESSAGES_LOG_SQL, CREATE_ABSA_SQL])
    url     = f"https://api.supabase.com/v1/projects/{project_ref}/database/query"
    headers = {"Authorization": f"Bearer {service_key}", "Content-Type": "application/json"}
    try:
        resp = requests.post(url, headers=headers, json={"query": all_sql}, timeout=30)
        if resp.status_code in (200, 201):
            print("Tables created via Management API")
            return True
        print(f"Management API {resp.status_code}: {resp.text[:200]}")
        return False
    except Exception as e:
        print(f"Management API error: {e}")
        return False

def seed_customers_from_csv():
    csv_path = f"{DATA_DIR}/churn_scores_final.csv"
    if not os.path.exists(csv_path):
        print(f"{csv_path} not found - skipping seed.")
        return 0
    df = pd.read_csv(csv_path)
    try:
        sb    = get_supabase()
        check = sb.table("customers").select("id", count="exact").execute()
        if check.count and check.count >= len(df):
            print(f"customers already has {check.count} rows - skipping seed")
            return check.count
    except Exception as e:
        print(f"customers table check failed (may not exist yet): {e}")  # PATCHED: was bare except: pass
    print(f"Seeding {len(df)} customers...")
    col_map = {
        'Last 4 Card Digits':'id','Segment':'segment','Recency':'recency',
        'Frequency':'frequency','Monetary':'monetary','Unique_Items':'unique_items',
        'Avg_Order_Val':'avg_order_val','Avg_Tip':'avg_tip','Discount_Used':'discount_used',
        'Visits_Nov':'visits_nov','Visits_Dec':'visits_dec','Visits_Jan':'visits_jan',
        'Days_Since_First':'days_since_first','Churn_Probability':'churn_probability',
        'Risk_Level':'risk_level','Tier':'tier',
    }
    df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
    df['id']               = df['id'].astype(str).str.strip()
    df['discount_offered'] = df['risk_level'].map(lambda r: f"{DISCOUNT_MAP.get(str(r), 15)}%")
    df = df.fillna(0)
    keep    = ['id','segment','recency','frequency','monetary','unique_items','avg_order_val',
               'avg_tip','discount_used','visits_nov','visits_dec','visits_jan',
               'days_since_first','churn_probability','risk_level','tier','discount_offered']
    df      = df[[c for c in keep if c in df.columns]]
    records = [
        {k: (int(v) if isinstance(v, np.integer) else float(v) if isinstance(v, np.floating) else v)
         for k, v in row.items()}
        for row in df.to_dict('records')
    ]
    sb     = get_supabase()
    seeded = 0
    for i in range(0, len(records), 500):
        try:
            sb.table("customers").upsert(records[i:i+500], on_conflict="id").execute()
            seeded += len(records[i:i+500])
        except Exception as e:
            print(f"Seed batch error: {e}")
    print(f"Seeded {seeded} customers")
    return seeded

# ─────────────────────────────────────────────
# APP SETUP  (lifespan replaces deprecated @app.on_event)
# ─────────────────────────────────────────────
@asynccontextmanager
async def lifespan(app_instance):
    global menu_df
    load_models()
    menu_df = load_menu()
    print(f"Menu loaded: {len(menu_df)} items")
    ensure_tables_via_api()
    seed_customers_from_csv()
    yield   # app runs here
    # (shutdown logic goes here if needed)

app = FastAPI(
    title       = "Campbell's AI Marketing API",
    description = "Churn Prediction · Customer Segmentation · ABSA · Personalized Messages",
    version     = "4.0.0",
    lifespan    = lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins     = ["*"],
    allow_credentials = True,
    allow_methods     = ["*"],
    allow_headers     = ["*"],
)

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────
def get_segment(recency, frequency, monetary) -> str:
    if 'kmeans' not in models:
        raise HTTPException(status_code=503, detail="Segmentation model not loaded")
    scaled  = models['scaler'].transform([[recency, frequency, monetary]])
    cluster = models['kmeans'].predict(scaled)[0]
    return models['cluster_map'].get(cluster, 'Unknown')

def get_churn_score(req: CustomerRequest, segment: str) -> dict:
    if req.frequency == 1:
        prob = 0.30 if req.recency < 14 else (0.65 if req.recency < 30 else 0.92)
        tier = "Rule-Based"
    else:
        if 'churn' not in models:
            raise HTTPException(status_code=503, detail="Churn model not loaded")
        seg_map  = {'Regular': 0, 'New': 1, 'Occasional': 2, 'Lost': 3}
        features = {
            'Recency'        : req.recency,
            'Frequency'      : req.frequency,
            'Monetary'       : req.monetary,
            'Unique_Items'   : req.unique_items or 1,
            'Avg_Order_Val'  : req.avg_order_val or req.monetary,
            'Avg_Tip'        : req.avg_tip or 0.0,
            'Discount_Used'  : req.discount_used or 0,
            'Visits_Nov'     : req.visits_nov or 0,
            'Visits_Dec'     : req.visits_dec or 0,
            'Visits_Jan'     : req.visits_jan or 0,
            'Days_Since_First': req.days_since_first or int(req.recency),
            'Segment_Code'   : seg_map.get(segment, 1)
        }
        X    = pd.DataFrame([features])[models['churn_feats']]
        prob = float(models['churn'].predict_proba(X)[0][1])
        tier = "XGBoost"
    risk = "Low" if prob < 0.33 else ("Medium" if prob < 0.66 else "High")
    return {"churn_probability": round(prob, 4), "risk_level": risk, "tier": tier}

OPINION_KEYWORDS = {
    'food'    : ['food','dish','meal','taste','flavor','delicious','bland','overcooked','fresh','portion'],
    'staff'   : ['waiter','waitress','server','staff','host','bartender','friendly','rude','helpful','attentive'],
    'service' : ['service','wait','slow','fast','quick','prompt','attentive','responsive'],
    'place'   : ['place','location','restaurant','spot','venue','seating','atmosphere'],
    'menu'    : ['menu','options','variety','selection','choice','specials'],
    'ambience': ['ambience','ambiance','atmosphere','decor','noise','cozy','loud','vibe','setting'],
    'price'   : ['price','expensive','cheap','value','worth','overpriced','affordable','cost']
}

def extract_opinion(review: str, aspect: str) -> str:
    keywords  = OPINION_KEYWORDS.get(aspect, [aspect])
    sentences = re.split(r'[.!?,;]', review)
    for sent in sentences:
        if any(kw in sent.lower() for kw in keywords):
            words = sent.strip().split()
            for i, word in enumerate(words):
                if word.lower().strip('.,!?') in keywords:
                    snippet = ' '.join(words[max(0,i-3):min(len(words),i+4)]).strip('.,!? ')
                    if snippet: return snippet
    return aspect

def run_absa(review: str) -> dict:
    if 'aspect' not in models:
        raise HTTPException(status_code=503, detail="ABSA model not loaded")
    X_r       = models['tfidf_asp'].transform([review])
    aspects   = list(models['mlb'].inverse_transform(models['aspect'].predict(X_r))[0])
    sentiments, opinions, triplets = [], [], []
    for asp in aspects:
        inp       = models['tfidf_sent'].transform([review + ' [ASPECT] ' + asp])
        sentiment = models['sentiment'].predict(inp)[0]
        opinion   = extract_opinion(review, asp)
        sentiments.append(sentiment)
        opinions.append(opinion)
        triplets.append({"aspect": asp, "opinion": opinion, "sentiment": sentiment})
    return {"aspects": aspects, "sentiments": sentiments, "opinions": opinions, "triplets": triplets}

def call_groq(prompt: str) -> str:
    headers = {
        "Content-Type" : "application/json",
        "Authorization": f"Bearer {GROQ_API_KEY}"
    }
    body = {
        "model"          : GROQ_MODEL,
        "max_tokens"     : 1000,
        "temperature"    : 0.7,
        "response_format": {"type": "json_object"},
        "messages"       : [
            {"role": "system", "content": "You are a marketing assistant for Campbell's Restaurant. Respond with valid JSON only."},
            {"role": "user",   "content": prompt}
        ]
    }
    resp = requests.post(
        "https://api.groq.com/openai/v1/chat/completions",
        headers=headers, json=body, timeout=30
    )
    resp.raise_for_status()
    return resp.json()['choices'][0]['message']['content']

def generate_message(req: MessageRequest) -> dict:
    discount = DISCOUNT_MAP.get(req.risk_level, 15)
    liked    = [a for a, s in zip(req.aspects or [], req.sentiments or []) if s == 'positive']
    disliked = [a for a, s in zip(req.aspects or [], req.sentiments or []) if s == 'negative']

    highlights = []
    if not menu_df.empty:
        food_pref = (req.food_preference or 'varied').lower()
        if food_pref != 'varied':
            matching = menu_df[menu_df['itemName'].str.lower().str.contains(food_pref, na=False)]
            if len(matching) > 0:
                highlights = matching.sample(min(2, len(matching)))[['itemName','itemPrice','Category']].to_dict('records')
        if len(highlights) < 3:
            extra = menu_df.sample(min(3 - len(highlights), len(menu_df)))[['itemName','itemPrice','Category']].to_dict('records')
            highlights.extend(extra)

    persona_notes = []
    tier         = req.spending_tier    or 'Standard'
    time_pref    = req.time_preference  or 'Evening'
    drink_food   = req.drink_vs_food    or 'Mixed'
    flight_lover = req.is_flight_lover  or False
    food_pref    = req.food_preference  or 'varied'
    modifier     = req.favorite_modifier

    if tier == 'Premium':         persona_notes.append("VIP Premium spender - treat them like royalty, use luxury language")
    elif tier == 'Economy':       persona_notes.append("Value-conscious - lead with the deal, make savings feel big")
    if time_pref == 'Evening':    persona_notes.append("Evening regular - reference their night-out vibe naturally")
    elif time_pref == 'Morning':  persona_notes.append("Morning person - brunch/breakfast items may appeal")
    elif time_pref == 'Mid-day':  persona_notes.append("Lunch crowd regular - highlight midday specials")
    if drink_food == 'Drinks':    persona_notes.append("Drinks-focused - highlight cocktail and drink flights")
    elif drink_food == 'Food':    persona_notes.append("Food-focused - highlight food items over drinks")
    if flight_lover:              persona_notes.append("LOVES flights - use flight/aviation puns (cleared for takeoff, fly back, boarding)")
    if modifier:                  persona_notes.append(f"Always orders with {modifier} - mention it by name")
    if food_pref != 'varied':     persona_notes.append(f"Loves {food_pref} dishes - recommend {food_pref} items from menu")
    if disliked:                  persona_notes.append(f"Previously unhappy with {', '.join(disliked)} - subtly acknowledge improvement")

    prompt = f"""You are a warm, creative marketing copywriter for Campbell's Restaurant.

CUSTOMER PROFILE:
- Segment              : {req.segment}
- Days since last visit: {int(req.recency)} days
- Total visits         : {req.frequency}
- Total spent          : ${req.monetary:.2f}
- Churn risk           : {req.risk_level}
- Spending tier        : {tier}
- Favorite items       : {req.favorite_items or []}
- Preferred time       : {time_pref}
- Preference type      : {drink_food} person
- Food loves           : {food_pref}
- Favorite modifier    : {modifier if modifier else 'not specified'}
- Flight lover         : {'Yes - use flight puns!' if flight_lover else 'Not particularly'}
- Liked aspects        : {liked if liked else 'unknown'}
- Disliked aspects     : {disliked if disliked else 'none noted'}

PERSONA NOTES (follow these carefully):
{chr(10).join(f'- {note}' for note in persona_notes)}

MENU HIGHLIGHTS (pick what fits the customer):
{json.dumps(highlights, indent=2)}

DISCOUNT TO OFFER: {discount}% off next visit

Return ONLY this JSON object:
{{
  "sms": "...",
  "email": {{"subject": "...", "body": "..."}},
  "app_notification": "..."
}}

WRITING RULES:
- SMS: max 160 chars, reference their specific preferences, punchy
- Email: warm, personal, 3-4 short paragraphs, mention their favorites + modifier if known
- App notification: max 80 chars, include 1 relevant emoji
- Reference their TIME preference naturally (evening rush, morning crowd etc.)
- Spending tier tone: Premium = VIP language, Economy = savings-first language
- If flight lover: use puns like 'cleared for takeoff', 'fly back', 'your table is boarding'
- Include the {discount}% discount clearly and naturally
- Tone by segment: Lost = urgent miss you, New = welcoming, Occasional = appreciate loyalty, Regular = VIP
- NEVER mention churn, AI, data analysis, or risk scores"""

    response = call_groq(prompt)
    try:
        clean      = re.sub(r'```json|```', '', response).strip()
        json_match = re.search(r'\{.*\}', clean, re.DOTALL)
        if json_match: clean = json_match.group(0)
        return json.loads(clean)
    except:
        return {"sms": response[:160], "email": {}, "app_notification": response[:80]}

# ─────────────────────────────────────────────
# SUPABASE WRITE HELPERS
# ─────────────────────────────────────────────
def upsert_customer_to_db(customer_id, segment, req, churn):
    if not customer_id: return
    sb = get_supabase()
    try:
        sb.table("customers").upsert({
            "id"               : str(customer_id),
            "segment"          : segment,
            "recency"          : float(req.recency),
            "frequency"        : int(req.frequency),
            "monetary"         : float(req.monetary),
            "unique_items"     : int(req.unique_items or 1),
            "avg_order_val"    : float(req.avg_order_val or req.monetary),
            "avg_tip"          : float(req.avg_tip or 0),
            "discount_used"    : int(req.discount_used or 0),
            "visits_nov"       : int(req.visits_nov or 0),
            "visits_dec"       : int(req.visits_dec or 0),
            "visits_jan"       : int(req.visits_jan or 0),
            "days_since_first" : int(req.days_since_first or req.recency),
            "churn_probability": float(churn['churn_probability']),
            "risk_level"       : churn['risk_level'],
            "tier"             : churn['tier'],
            "discount_offered" : f"{DISCOUNT_MAP.get(churn['risk_level'], 15)}%",
            "updated_at"       : datetime.now(timezone.utc).isoformat(),
        }, on_conflict="id").execute()
    except Exception as e:
        print(f"Customer upsert failed: {e}")

def log_message_to_db(customer_id, req: MessageRequest, messages: dict, absa):
    sb    = get_supabase()
    email = messages.get("email", {})
    try:
        sb.table("messages_log").insert({
            "customer_id"     : str(customer_id) if customer_id else None,
            "segment"         : req.segment,
            "risk_level"      : req.risk_level,
            "discount_offered": f"{DISCOUNT_MAP.get(req.risk_level, 15)}%",
            "spending_tier"   : req.spending_tier,
            "time_preference" : req.time_preference,
            "sms"             : messages.get("sms", ""),
            "email_subject"   : email.get("subject", "") if isinstance(email, dict) else "",
            "email_body"      : email.get("body", "")    if isinstance(email, dict) else "",
            "app_notification": messages.get("app_notification", ""),
            # Pass raw list — Supabase handles JSONB serialisation
            "aspects"         : absa['aspects']    if absa else [],
            "sentiments"      : absa['sentiments'] if absa else [],
        }).execute()
    except Exception as e:
        print(f"Message log failed: {e}")

def log_absa_to_db(customer_id, review: str, absa: dict):
    sb = get_supabase()
    try:
        sb.table("absa_predictions").insert({
            "customer_id"    : str(customer_id) if customer_id else None,
            "review"         : review,
            "aspects"        : json.dumps(absa['aspects']),
            "sentiments"     : json.dumps(absa['sentiments']),
            "feature_opinion": json.dumps(absa['opinions']),
            "opinions"       : json.dumps(absa['opinions']),
            "triplets"       : absa['triplets'],   # JSONB — pass raw list
        }).execute()
    except Exception as e:
        print(f"ABSA log failed: {e}")

# ─────────────────────────────────────────────
# EMAIL HELPERS
# ─────────────────────────────────────────────
REAL_EMAIL_MAP = {
    "342.0" : "gamingoutclassed@gmail.com",
    "6456.0": "murtazaworks0@gmail.com",
    "7383.0": "222134@students.au.edu.pk",
    "3066.0": "gamingoutclassed@gmail.com",
    "5650.0": "murtazaworks0@gmail.com",
    "3610.0": "222134@students.au.edu.pk",
    "4861.0": "gamingoutclassed@gmail.com",
    "4532.0": "murtazaworks0@gmail.com",
    "1563.0": "222134@students.au.edu.pk",
    "8753.0": "gamingoutclassed@gmail.com",
}

def get_customer_email(customer_id: str) -> str:
    cid = str(customer_id).strip()
    if cid in REAL_EMAIL_MAP:
        return REAL_EMAIL_MAP[cid]
    try:
        sb   = get_supabase()
        resp = sb.table("customers").select("email").eq("id", cid).limit(1).execute()
        if resp.data and resp.data[0].get("email"):
            return resp.data[0]["email"]
    except:
        pass
    clean_id = cid.replace(".", "").replace(" ", "")
    return f"customer_{clean_id}@demo.campbells-restaurant.com"


def send_email_via_resend(
    to_email: str,
    subject: str,
    body: str,
    customer_id: str = None,
    segment: str     = None,
    risk_level: str  = None,
    discount: str    = None
) -> dict:
    if resend is None:
        return {"success": False, "error": "Resend not configured", "to": to_email}

    segment_color = {
        "Lost"      : "#E05252",
        "New"       : "#4A9EFF",
        "Occasional": "#F5A623",
        "Regular"   : "#2DD4A0",
    }.get(segment, "#FF6B35")

    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1.0">
      <title>{subject}</title>
    </head>
    <body style="margin:0; padding:0; background-color:#0A0A0F; font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;">
      <div style="background: linear-gradient(135deg, #FF6B35 0%, #FF8C5A 100%); padding: 32px 40px; text-align: center;">
        <p style="margin:0; color:white; font-size:13px; letter-spacing:3px; text-transform:uppercase; opacity:0.85;">
          ✈ CLEARED FOR TAKEOFF
        </p>
        <h1 style="margin:8px 0 0 0; color:white; font-size:32px; font-weight:700; letter-spacing:-0.5px;">
          Campbell's
        </h1>
        <p style="margin:4px 0 0 0; color:rgba(255,255,255,0.8); font-size:13px; letter-spacing:2px; text-transform:uppercase;">
          AI Marketing Hub
        </p>
      </div>
      <div style="background:#12121A; padding:40px; max-width:600px; margin:0 auto;">
        <div style="margin-bottom:24px;">
          <span style="background:{segment_color}22; color:{segment_color}; border:1px solid {segment_color}44;
                       padding:4px 12px; border-radius:20px; font-size:12px; font-weight:600;
                       text-transform:uppercase; letter-spacing:1px;">
            {segment or 'Customer'} Member
          </span>
          {f'<span style="background:#E0525222; color:#E05252; border:1px solid #E0525244; padding:4px 12px; border-radius:20px; font-size:12px; font-weight:600; text-transform:uppercase; letter-spacing:1px; margin-left:8px;">Special Offer: {discount}</span>' if discount else ''}
        </div>
        <div style="color:#E0E0E8; font-size:15px; line-height:1.8; white-space:pre-line;">
{body}
        </div>
        <div style="text-align:center; margin:36px 0;">
          <a href="https://campbells-ai-hub.vercel.app"
             style="background: linear-gradient(135deg, #FF6B35, #FF8C5A);
                    color:white; text-decoration:none; padding:14px 36px;
                    border-radius:8px; font-weight:600; font-size:15px;
                    display:inline-block; letter-spacing:0.3px;">
            ✈ Fly Back to Campbell's
          </a>
        </div>
        <div style="border-top:1px solid rgba(255,255,255,0.06); margin:32px 0;"></div>
        <div style="text-align:center; color:#4A4A5E; font-size:12px; line-height:1.6;">
          <p style="margin:0;">Campbell's Restaurant · Flight-Themed Dining Experience</p>
          <p style="margin:4px 0 0 0;">
            This message was generated by Campbell's AI Marketing Hub.
            <br>You are receiving this because you are a valued guest.
          </p>
          {f'<p style="margin:8px 0 0 0; color:#2A2A3E; font-size:10px;">Customer ID: {customer_id} | Sent via Campbell\'s AI System</p>' if customer_id else ''}
        </div>
      </div>
    </body>
    </html>
    """

    try:
        response = resend.Emails.send({
            "from"   : "Campbell's AI Hub <onboarding@resend.dev>",
            "to"     : [to_email],
            "subject": subject,
            "html"   : html_body,
        })
        print(f"Email sent to {to_email} | ID: {response.get('id', 'unknown')}")
        return {"success": True, "email_id": response.get("id"), "to": to_email}
    except Exception as e:
        print(f"Email send failed to {to_email}: {e}")
        return {"success": False, "error": str(e), "to": to_email}


def update_email_sent_status(customer_id: str, email: str, success: bool):
    if not customer_id:
        return
    sb = get_supabase()
    try:
        sb.table("customers").update({
            "email"         : email,
            "email_sent"    : success,
            "email_sent_at" : datetime.now(timezone.utc).isoformat(),
        }).eq("id", str(customer_id)).execute()
    except Exception as e:
        print(f"Email status update failed: {e}")

# ─────────────────────────────────────────────
# CORE ENDPOINTS
# ─────────────────────────────────────────────
@app.get("/")
def health_check():
    return {
        "status"       : "ok",
        "service"      : "Campbell's AI Marketing API",
        "version"      : "4.0.0",
        "models_loaded": list(models.keys()),
        "supabase"     : SUPABASE_URL != "YOUR_SUPABASE_URL_HERE",
        "email"        : resend is not None,
    }

@app.post("/api/segment")
def segment_customer(req: CustomerRequest):
    segment = get_segment(req.recency, req.frequency, req.monetary)
    return {"segment": segment, "recency": req.recency,
            "frequency": req.frequency, "monetary": req.monetary}

@app.post("/api/churn")
def predict_churn(req: CustomerRequest):
    segment = get_segment(req.recency, req.frequency, req.monetary)
    result  = get_churn_score(req, segment)
    return {
        "segment"          : segment,
        "churn_probability": result['churn_probability'],
        "risk_level"       : result['risk_level'],
        "tier"             : result['tier'],
        "discount_to_offer": f"{DISCOUNT_MAP.get(result['risk_level'], 15)}%"
    }

@app.post("/api/sentiment")
def analyze_sentiment(req: SentimentRequest):
    result = run_absa(req.review)
    return {
        "review"    : req.review,
        "aspects"   : result['aspects'],
        "sentiments": result['sentiments'],
        "opinions"  : result['opinions'],
        "triplets"  : result['triplets']
    }

@app.post("/api/generate-message")
def generate_personalized_message(req: MessageRequest):
    messages = generate_message(req)
    log_message_to_db(req.customer_id, req, messages, None)
    return {
        "customer_id"     : req.customer_id,
        "segment"         : req.segment,
        "risk_level"      : req.risk_level,
        "discount_offered": f"{DISCOUNT_MAP.get(req.risk_level, 15)}%",
        "messages"        : messages
    }

@app.post("/api/full-pipeline")
def full_pipeline(req: FullPipelineRequest):
    segment   = get_segment(req.recency, req.frequency, req.monetary)
    churn_req = CustomerRequest(
        recency         = req.recency,
        frequency       = req.frequency,
        monetary        = req.monetary,
        unique_items    = req.unique_items,
        avg_order_val   = req.avg_order_val,
        avg_tip         = req.avg_tip,
        discount_used   = req.discount_used,
        visits_nov      = req.visits_nov,
        visits_dec      = req.visits_dec,
        visits_jan      = req.visits_jan,
        days_since_first= req.days_since_first,
        favorite_items  = req.favorite_items,
    )
    churn = get_churn_score(churn_req, segment)

    absa = None
    if req.review:
        absa = run_absa(req.review)
        log_absa_to_db(req.customer_id, req.review, absa)

    # Resolve Section 5 profile: request body > Supabase DB > defaults
    spending_tier     = req.spending_tier
    time_preference   = req.time_preference
    food_preference   = req.food_preference
    drink_vs_food     = req.drink_vs_food
    favorite_modifier = req.favorite_modifier
    is_flight_lover   = req.is_flight_lover
    fav_items         = req.favorite_items or []

    if req.customer_id:
        try:
            sb   = get_supabase()
            resp = sb.table("customers").select(
                "spending_tier,time_preference,food_preference,drink_vs_food,"
                "favorite_modifier,is_flight_lover,favorite_items"
            ).eq("id", str(req.customer_id)).limit(1).execute()
            if resp.data:
                row = resp.data[0]
                if spending_tier     is None: spending_tier     = row.get("spending_tier")
                if time_preference   is None: time_preference   = row.get("time_preference")
                if food_preference   is None: food_preference   = row.get("food_preference")
                if drink_vs_food     is None: drink_vs_food     = row.get("drink_vs_food")
                if favorite_modifier is None: favorite_modifier = row.get("favorite_modifier")
                if is_flight_lover   is None: is_flight_lover   = bool(row.get("is_flight_lover", False))
                if not fav_items:
                    stored = row.get("favorite_items")
                    if stored:
                        try:
                            fav_items = ast.literal_eval(stored) if isinstance(stored, str) else stored
                        except:
                            pass
        except Exception as e:
            print(f"Profile fetch error: {e}")

    spending_tier    = spending_tier    or 'Standard'
    time_preference  = time_preference  or 'Evening'
    food_preference  = food_preference  or 'varied'
    drink_vs_food    = drink_vs_food    or 'Mixed'
    is_flight_lover  = is_flight_lover  or False

    msg_req = MessageRequest(
        customer_id      = req.customer_id,
        segment          = segment,
        recency          = req.recency,
        frequency        = req.frequency,
        monetary         = req.monetary,
        risk_level       = churn['risk_level'],
        churn_probability= churn['churn_probability'],
        favorite_items   = fav_items,
        aspects          = absa['aspects']    if absa else [],
        sentiments       = absa['sentiments'] if absa else [],
        spending_tier    = spending_tier,
        time_preference  = time_preference,
        food_preference  = food_preference,
        drink_vs_food    = drink_vs_food,
        favorite_modifier= favorite_modifier,
        is_flight_lover  = is_flight_lover,
    )
    messages = generate_message(msg_req)

    upsert_customer_to_db(req.customer_id, segment, req, churn)
    log_message_to_db(req.customer_id, msg_req, messages, absa)

    # PATCHED: Only send to real mapped emails — skip generated demo addresses
    email_result   = {"success": False, "skipped": True, "reason": "no real email on file"}
    customer_email = get_customer_email(req.customer_id) if req.customer_id else None

    is_real_email = (
        customer_email is not None
        and "demo.campbells-restaurant.com" not in customer_email
    )

    if is_real_email:
        email_data = messages.get("email", {})
        subject    = email_data.get("subject", "A message from Campbell's") if isinstance(email_data, dict) else "A message from Campbell's"
        body       = email_data.get("body", "")                              if isinstance(email_data, dict) else str(email_data)
        discount   = f"{DISCOUNT_MAP.get(churn['risk_level'], 15)}%"

        email_result = send_email_via_resend(
            to_email    = customer_email,
            subject     = subject,
            body        = body,
            customer_id = req.customer_id,
            segment     = segment,
            risk_level  = churn['risk_level'],
            discount    = discount,
        )
        update_email_sent_status(req.customer_id, customer_email, email_result.get("success", False))

    return {
        "segment"          : segment,
        "churn_probability": churn['churn_probability'],
        "risk_level"       : churn['risk_level'],
        "tier"             : churn['tier'],
        "discount_offered" : f"{DISCOUNT_MAP.get(churn['risk_level'], 15)}%",
        "profile"          : {
            "spending_tier"  : spending_tier,
            "time_preference": time_preference,
            "food_preference": food_preference,
            "drink_vs_food"  : drink_vs_food,
            "is_flight_lover": is_flight_lover,
        },
        "absa"        : absa,
        "messages"    : messages,
        "email_result": email_result,
    }

# ─────────────────────────────────────────────
# BATCH SEND ENDPOINT
# ─────────────────────────────────────────────
@app.post("/api/send-reengagement-batch")
async def send_reengagement_batch(limit: int = Query(10, ge=1, le=50)):
    """
    Fetch top N high-risk customers from Supabase,
    generate personalized messages, and send emails via Resend.
    Only sends to customers with real mapped emails — demo addresses are skipped.
    """
    sb = get_supabase()
    try:
        resp = sb.table("customers").select("*") \
                 .eq("risk_level", "High") \
                 .order("churn_probability", desc=True) \
                 .limit(limit).execute()
        customers = resp.data
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

    results = []
    for customer in customers:
        cid   = customer.get("id")
        email = get_customer_email(cid)

        msg_req = MessageRequest(
            customer_id      = cid,
            segment          = customer.get("segment", "Lost"),
            recency          = float(customer.get("recency", 60)),
            frequency        = int(customer.get("frequency", 1)),
            monetary         = float(customer.get("monetary", 0)),
            risk_level       = customer.get("risk_level", "High"),
            churn_probability= float(customer.get("churn_probability", 0.9)),
            spending_tier    = customer.get("spending_tier", "Standard"),
            time_preference  = customer.get("time_preference", "Evening"),
            food_preference  = customer.get("food_preference", "varied"),
            drink_vs_food    = customer.get("drink_vs_food", "Mixed"),
            favorite_modifier= customer.get("favorite_modifier"),
            is_flight_lover  = bool(customer.get("is_flight_lover", False)),
        )

        try:
            stored = customer.get("favorite_items", "[]")
            msg_req.favorite_items = ast.literal_eval(stored) if isinstance(stored, str) else []
        except:
            msg_req.favorite_items = []

        # PATCHED: 800ms delay between Groq calls — prevents free-tier 429 rate limit errors
        messages = generate_message(msg_req)
        log_message_to_db(cid, msg_req, messages, None)

        email_data = messages.get("email", {})
        subject    = email_data.get("subject", "A message from Campbell's") if isinstance(email_data, dict) else "A message from Campbell's"
        body       = email_data.get("body", "") if isinstance(email_data, dict) else ""

        # PATCHED: Only send to real mapped emails — skip generated demo addresses
        is_real_email = (
            email is not None
            and "demo.campbells-restaurant.com" not in email
        )

        if is_real_email:
            email_result = send_email_via_resend(
                to_email    = email,
                subject     = subject,
                body        = body,
                customer_id = cid,
                segment     = customer.get("segment"),
                risk_level  = customer.get("risk_level"),
                discount    = f"{DISCOUNT_MAP.get(customer.get('risk_level','High'), 20)}%"
            )
            update_email_sent_status(cid, email, email_result.get("success", False))
        else:
            email_result = {"success": False, "skipped": True, "reason": "demo address"}

        results.append({
            "customer_id"  : cid,
            "email"        : email,
            "real_email"   : is_real_email,
            "segment"      : customer.get("segment"),
            "sms"          : messages.get("sms", ""),
            "email_subject": subject,
            "email_sent"   : email_result.get("success", False),
            "email_id"     : email_result.get("email_id"),
        })

        time.sleep(0.8)   # rate-limit buffer between Groq calls

    sent_count = sum(1 for r in results if r["email_sent"])
    return {
        "total_processed": len(results),
        "emails_sent"    : sent_count,
        "results"        : results
    }

# ─────────────────────────────────────────────
# DASHBOARD + DATA ENDPOINTS
# ─────────────────────────────────────────────
@app.get("/api/dashboard")
def get_dashboard_stats():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "segment, risk_level, recency, frequency, monetary, churn_probability"
        ).limit(10000).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        print(f"Supabase fallback: {e}")
        try:
            df = pd.read_csv(f"{DATA_DIR}/churn_scores_final.csv")
            df = df.rename(columns={
                'Segment':'segment','Risk_Level':'risk_level','Recency':'recency',
                'Frequency':'frequency','Monetary':'monetary','Churn_Probability':'churn_probability'
            })
        except:
            raise HTTPException(status_code=503, detail="No data available")

    seg_counts     = df['segment'].value_counts().to_dict()    if 'segment'    in df.columns else {}
    risk_counts    = df['risk_level'].value_counts().to_dict() if 'risk_level' in df.columns else {}
    churn_rate     = round(float(df['churn_probability'].mean()) * 100, 2) if 'churn_probability' in df.columns else 0
    top_risk       = (
        df[df['risk_level'].astype(str) == 'High']
        .sort_values('churn_probability', ascending=False)
        .head(20)[['segment','recency','frequency','monetary','churn_probability','risk_level']]
        .to_dict('records')
    ) if 'risk_level' in df.columns else []
    rfm_by_segment = df.groupby('segment')[['recency','frequency','monetary']].mean().round(2).to_dict('index') if 'segment' in df.columns else {}
    messages_sent  = 0
    try:
        msg_resp      = sb.table("messages_log").select("id", count="exact").execute()
        messages_sent = msg_resp.count or 0
    except: pass
    return {
        "total_customers": int(len(df)),
        "segment_counts" : seg_counts,
        "risk_counts"    : risk_counts,
        "churn_rate_pct" : churn_rate,
        "top_at_risk"    : top_risk,
        "rfm_by_segment" : rfm_by_segment,
        "messages_sent"  : messages_sent,
        "data_source"    : "supabase"
    }

@app.get("/api/customers")
def list_customers(
    segment   : Optional[str] = Query(None),
    risk_level: Optional[str] = Query(None),
    page      : int           = Query(1,  ge=1),
    page_size : int = Query(50, ge=1, le=2000)
):
    sb     = get_supabase()
    offset = (page - 1) * page_size
    query  = sb.table("customers").select("*", count="exact")
    if segment:    query = query.eq("segment",    segment)
    if risk_level: query = query.eq("risk_level", risk_level)
    try:
        resp = query.range(offset, offset + page_size - 1).execute()
        return {
            "page"       : page,
            "page_size"  : page_size,
            "total"      : resp.count,
            "total_pages": -(-resp.count // page_size) if resp.count else 0,
            "customers"  : resp.data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

@app.get("/api/messages-log")
def get_messages_log(
    limit  : int           = Query(50, ge=1, le=200),
    segment: Optional[str] = Query(None)
):
    sb    = get_supabase()
    query = sb.table("messages_log").select("*").order("created_at", desc=True)
    if segment: query = query.eq("segment", segment)
    try:
        resp = query.limit(limit).execute()
        return {"count": len(resp.data), "messages": resp.data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

@app.get("/api/customer-profile/{customer_id}")
def get_customer_profile(customer_id: str):
    """Full enriched profile for one customer including all Section 5 fields."""
    sb = get_supabase()
    try:
        resp = sb.table("customers").select("*").eq("id", customer_id).execute()
        if not resp.data:
            raise HTTPException(status_code=404, detail=f"Customer {customer_id} not found")
        return resp.data[0]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB error: {e}")

# ─────────────────────────────────────────────
# ANALYSIS ENDPOINTS
# ─────────────────────────────────────────────
@app.get("/api/analysis/kpis")
def get_kpis():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "segment, risk_level, churn_probability, monetary, frequency"
        ).limit(10000).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    messages_sent = 0
    try:
        msg_resp      = sb.table("messages_log").select("id", count="exact").execute()
        messages_sent = msg_resp.count or 0
    except: pass

    urgent_count = 0
    try:
        urgent_resp  = sb.table("customers").select("id", count="exact") \
                         .eq("risk_level", "High").gte("recency", 60).execute()
        urgent_count = urgent_resp.count or 0
    except: pass

    return {
        "total_customers"  : len(df),
        "churn_rate_pct"   : round(float(df['churn_probability'].mean()) * 100, 2),
        "high_risk_count"  : int((df['risk_level'].astype(str) == 'High').sum()),
        "medium_risk_count": int((df['risk_level'].astype(str) == 'Medium').sum()),
        "low_risk_count"   : int((df['risk_level'].astype(str) == 'Low').sum()),
        "avg_spend"        : round(float(df['monetary'].mean()), 2),
        "total_revenue"    : round(float(df['monetary'].sum()), 2),
        "messages_sent"    : messages_sent,
        "top_segment"      : df['segment'].value_counts().idxmax(),
        "avg_visits"       : round(float(df['frequency'].mean()), 2),
        "urgent_reachout"  : urgent_count,
    }

@app.get("/api/analysis/rfm")
def get_rfm_data():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "id, segment, recency, frequency, monetary, churn_probability, risk_level"
        ).limit(500).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")
    seg_avg = df.groupby('segment')[['recency','frequency','monetary']].mean().round(2).to_dict('index')
    return {
        "scatter_data"    : df[['id','segment','recency','frequency','monetary',
                                'churn_probability','risk_level']].to_dict('records'),
        "segment_averages": seg_avg,
        "total_points"    : len(df)
    }

@app.get("/api/analysis/churn-distribution")
def get_churn_distribution():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "segment, churn_probability, risk_level"
        ).limit(10000).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    df['prob_bucket'] = pd.cut(
        df['churn_probability'],
        bins=[0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0],
        labels=['0-10%','10-20%','20-30%','30-40%','40-50%',
                '50-60%','60-70%','70-80%','80-90%','90-100%']
    )
    return {
        "histogram"        : df['prob_bucket'].value_counts().sort_index().to_dict(),
        "churn_by_segment" : df.groupby('segment')['churn_probability'].mean().round(4).to_dict(),
        "risk_by_segment"  : df.groupby(['segment','risk_level']).size().reset_index(name='count').to_dict('records'),
        "overall_avg_churn": round(float(df['churn_probability'].mean()), 4)
    }

@app.get("/api/analysis/monthly-visits")
def get_monthly_visits():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "segment, visits_nov, visits_dec, visits_jan, frequency"
        ).limit(10000).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    monthly_totals = {
        "November": int(df['visits_nov'].sum()),
        "December": int(df['visits_dec'].sum()),
        "January" : int(df['visits_jan'].sum()),
    }
    monthly_by_segment = {}
    for seg in df['segment'].unique():
        seg_df = df[df['segment'] == seg]
        monthly_by_segment[seg] = {
            "November": round(float(seg_df['visits_nov'].mean()), 2),
            "December": round(float(seg_df['visits_dec'].mean()), 2),
            "January" : round(float(seg_df['visits_jan'].mean()), 2),
        }
    freq_dist = df['frequency'].value_counts().sort_index()
    return {
        "monthly_totals"        : monthly_totals,
        "monthly_by_segment"    : monthly_by_segment,
        "frequency_distribution": {
            "1 visit"  : int(freq_dist.get(1, 0)),
            "2 visits" : int(freq_dist.get(2, 0)),
            "3 visits" : int(freq_dist.get(3, 0)),
            "4 visits" : int(freq_dist.get(4, 0)),
            "5+ visits": int(freq_dist[freq_dist.index >= 5].sum()),
        }
    }

@app.get("/api/analysis/revenue")
def get_revenue_analysis():
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "id, segment, monetary, avg_order_val, frequency, risk_level, discount_used"
        ).limit(10000).execute()
        rows = resp.data
        if not rows: raise ValueError("No data")
        df = pd.DataFrame(rows)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    revenue_by_segment = df.groupby('segment').agg(
        total_revenue  = ('monetary',      'sum'),
        avg_spend      = ('monetary',      'mean'),
        avg_order_val  = ('avg_order_val', 'mean'),
        customer_count = ('id',            'count'),
        avg_visits     = ('frequency',     'mean'),
    ).round(2).to_dict('index')

    df['spend_bucket'] = pd.cut(
        df['monetary'],
        bins=[0, 50, 100, 200, 500, 1000, 99999],
        labels=['$0-50','$50-100','$100-200','$200-500','$500-1000','$1000+']
    )
    return {
        "revenue_by_segment": revenue_by_segment,
        "spend_distribution": df['spend_bucket'].value_counts().sort_index().to_dict(),
        "top_spenders"      : df.nlargest(10, 'monetary')[['id','segment','monetary','frequency','risk_level']].to_dict('records'),
        "total_revenue"     : round(float(df['monetary'].sum()), 2),
        "avg_spend"         : round(float(df['monetary'].mean()), 2),
        "discount_users"    : int((df['discount_used'] > 0).sum()),
        "non_discount_users": int((df['discount_used'] == 0).sum()),
    }

@app.get("/api/analysis/sentiment-breakdown")
def get_sentiment_breakdown():
    sb = get_supabase()
    try:
        resp = sb.table("absa_predictions").select(
            "review, aspects, sentiments, feature_opinion"
        ).limit(10000).execute()
        rows = resp.data
        if not rows:
            raise HTTPException(status_code=404, detail="No ABSA data. Import absa_predictions.csv first.")
        df = pd.DataFrame(rows)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    def safe_parse(s):
        if isinstance(s, list): return s
        try:    return ast.literal_eval(str(s))
        except: return []

    pairs = []
    for _, row in df.iterrows():
        aspects    = safe_parse(row.get('aspects',    '[]'))
        sentiments = safe_parse(row.get('sentiments', '[]'))
        for asp, sent in zip(aspects, sentiments):
            pairs.append({'aspect': asp, 'sentiment': sent})

    if not pairs:
        return {"aspect_frequency": {}, "sentiment_by_aspect": {},
                "overall_sentiment": {}, "sample_reviews": [], "total_reviews": len(df)}

    pairs_df            = pd.DataFrame(pairs)
    aspect_frequency    = pairs_df['aspect'].value_counts().to_dict()
    sentiment_by_aspect = {
        asp: pairs_df[pairs_df['aspect'] == asp]['sentiment'].value_counts().to_dict()
        for asp in pairs_df['aspect'].unique()
    }
    return {
        "aspect_frequency"     : aspect_frequency,
        "sentiment_by_aspect"  : sentiment_by_aspect,
        "overall_sentiment"    : pairs_df['sentiment'].value_counts().to_dict(),
        "sample_reviews"       : df[['review','aspects','sentiments']].head(10).to_dict('records'),
        "total_reviews"        : len(df),
        "total_aspect_mentions": len(pairs_df)
    }

@app.get("/api/analysis/profiles")
def get_profile_breakdown():
    """
    Spending tier / time preference / food preference / flight lover breakdowns.
    Powers the Customer Profiling charts in the Analysis page.
    """
    sb = get_supabase()
    try:
        resp = sb.table("customers").select(
            "spending_tier, time_preference, food_preference, drink_vs_food, is_flight_lover, segment"
        ).limit(10000).execute()
        rows = resp.data
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"DB error: {e}")

    if not rows:
        raise HTTPException(status_code=404, detail="No profile data found")

    df = pd.DataFrame(rows)

    has_profiles = (
        "spending_tier" in df.columns
        and df["spending_tier"].notna().any()
        and (df["spending_tier"] != "Standard").any()
    )
    if not has_profiles:
        return {
            "message"         : "Profile columns exist but Section 5 data not yet imported. Run profiles_update_only.csv import first.",
            "spending_tiers"  : {},
            "time_preferences": {},
            "food_preferences": {},
            "drink_vs_food"   : {},
            "flight_lovers"   : 0,
            "total_profiled"  : 0,
        }

    def vc(col):
        return df[col].value_counts().to_dict() if col in df.columns else {}

    flight_lovers = 0
    if "is_flight_lover" in df.columns:
        flight_lovers = int(df["is_flight_lover"].apply(
            lambda x: bool(x) if x is not None else False
        ).sum())

    profile_by_segment = {}
    if "segment" in df.columns and "spending_tier" in df.columns:
        profile_by_segment = df.groupby("segment")["spending_tier"].value_counts().to_dict()

    return {
        "spending_tiers"    : vc("spending_tier"),
        "time_preferences"  : vc("time_preference"),
        "food_preferences"  : vc("food_preference"),
        "drink_vs_food"     : vc("drink_vs_food"),
        "flight_lovers"     : flight_lovers,
        "profile_by_segment": profile_by_segment,
        "total_profiled"    : len(df),
    }
